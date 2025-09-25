#!/usr/bin/env python3
"""
Performance benchmarking script for VAITP-Auditor GUI executables.

This script performs comprehensive performance testing including:
- Startup time analysis
- Memory usage profiling
- File loading performance
- UI responsiveness testing
- Resource utilization monitoring
"""

import os
import sys
import subprocess
import platform
import time
import psutil
import json
import tempfile
import threading
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import argparse
import statistics


class PerformanceBenchmark:
    """Performance benchmarking for executables."""
    
    def __init__(self, executable_path: Path):
        self.executable_path = Path(executable_path)
        self.system = platform.system().lower()
        self.results = {
            "platform": {
                "system": platform.system(),
                "release": platform.release(),
                "machine": platform.machine(),
                "processor": platform.processor(),
                "cpu_count": psutil.cpu_count(),
                "memory_total": psutil.virtual_memory().total / (1024**3),  # GB
            },
            "benchmarks": {},
            "timestamp": time.time()
        }
    
    def get_executable_command(self, args: List[str] = None) -> List[str]:
        """Get the command to run the executable."""
        if args is None:
            args = []
        
        if self.system == 'darwin' and self.executable_path.suffix == '.app':
            # For .app bundles, find the actual executable
            macos_dir = self.executable_path / 'Contents' / 'MacOS'
            if macos_dir.exists():
                executables = [f for f in macos_dir.iterdir() if f.is_file() and os.access(f, os.X_OK)]
                if executables:
                    return [str(executables[0])] + args
            raise RuntimeError("No executable found in app bundle")
        else:
            return [str(self.executable_path)] + args
    
    def benchmark_startup_time(self, iterations: int = 10) -> Dict:
        """Benchmark startup time with multiple iterations."""
        print(f"🚀 Benchmarking startup time ({iterations} iterations)...")
        
        startup_times = []
        cmd = self.get_executable_command(['--help'])
        
        for i in range(iterations):
            try:
                start_time = time.time()
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
                end_time = time.time()
                
                if result.returncode == 0:
                    startup_time = end_time - start_time
                    startup_times.append(startup_time)
                    print(f"  Iteration {i+1}: {startup_time:.3f}s")
                else:
                    print(f"  Iteration {i+1}: Failed (exit code {result.returncode})")
                    
            except subprocess.TimeoutExpired:
                print(f"  Iteration {i+1}: Timeout")
            except Exception as e:
                print(f"  Iteration {i+1}: Error - {e}")
        
        if startup_times:
            results = {
                "iterations": len(startup_times),
                "min_time": min(startup_times),
                "max_time": max(startup_times),
                "mean_time": statistics.mean(startup_times),
                "median_time": statistics.median(startup_times),
                "std_dev": statistics.stdev(startup_times) if len(startup_times) > 1 else 0,
                "raw_times": startup_times
            }
            
            print(f"  📊 Results:")
            print(f"    Min: {results['min_time']:.3f}s")
            print(f"    Max: {results['max_time']:.3f}s")
            print(f"    Mean: {results['mean_time']:.3f}s")
            print(f"    Median: {results['median_time']:.3f}s")
            print(f"    Std Dev: {results['std_dev']:.3f}s")
            
            return results
        else:
            return {"error": "No successful startup measurements"}
    
    def benchmark_memory_usage(self, duration: int = 10) -> Dict:
        """Benchmark memory usage over time."""
        print(f"🧠 Benchmarking memory usage ({duration}s monitoring)...")
        
        cmd = self.get_executable_command()
        memory_samples = []
        cpu_samples = []
        
        try:
            # Start the process
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            
            # Give it time to start
            time.sleep(1)
            
            try:
                ps_process = psutil.Process(process.pid)
                start_time = time.time()
                
                # Monitor for specified duration
                while time.time() - start_time < duration:
                    try:
                        memory_info = ps_process.memory_info()
                        cpu_percent = ps_process.cpu_percent()
                        
                        memory_mb = memory_info.rss / (1024 * 1024)
                        memory_samples.append(memory_mb)
                        cpu_samples.append(cpu_percent)
                        
                        time.sleep(0.5)  # Sample every 500ms
                        
                    except psutil.NoSuchProcess:
                        print("  Process exited during monitoring")
                        break
                
                # Terminate the process
                process.terminate()
                process.wait(timeout=5)
                
            except psutil.NoSuchProcess:
                print("  Process exited quickly")
            
        except Exception as e:
            print(f"  Error during memory benchmarking: {e}")
            return {"error": str(e)}
        
        if memory_samples:
            results = {
                "samples": len(memory_samples),
                "duration": duration,
                "memory": {
                    "min_mb": min(memory_samples),
                    "max_mb": max(memory_samples),
                    "mean_mb": statistics.mean(memory_samples),
                    "median_mb": statistics.median(memory_samples),
                    "std_dev_mb": statistics.stdev(memory_samples) if len(memory_samples) > 1 else 0,
                },
                "cpu": {
                    "min_percent": min(cpu_samples) if cpu_samples else 0,
                    "max_percent": max(cpu_samples) if cpu_samples else 0,
                    "mean_percent": statistics.mean(cpu_samples) if cpu_samples else 0,
                }
            }
            
            print(f"  📊 Memory Results:")
            print(f"    Min: {results['memory']['min_mb']:.1f} MB")
            print(f"    Max: {results['memory']['max_mb']:.1f} MB")
            print(f"    Mean: {results['memory']['mean_mb']:.1f} MB")
            print(f"    Median: {results['memory']['median_mb']:.1f} MB")
            
            print(f"  📊 CPU Results:")
            print(f"    Mean: {results['cpu']['mean_percent']:.1f}%")
            print(f"    Max: {results['cpu']['max_percent']:.1f}%")
            
            return results
        else:
            return {"error": "No memory samples collected"}
    
    def benchmark_file_loading(self) -> Dict:
        """Benchmark file loading performance with test files."""
        print("📁 Benchmarking file loading performance...")
        
        # Create test files
        test_files = self.create_test_files()
        loading_times = {}
        
        for file_type, file_path in test_files.items():
            print(f"  Testing {file_type} file: {file_path.name}")
            
            # For GUI applications, we can't easily test file loading directly
            # Instead, we'll test startup time with file arguments
            cmd = self.get_executable_command([str(file_path)])
            
            try:
                start_time = time.time()
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
                end_time = time.time()
                
                loading_time = end_time - start_time
                loading_times[file_type] = {
                    "time": loading_time,
                    "success": result.returncode == 0,
                    "file_size": file_path.stat().st_size
                }
                
                print(f"    Time: {loading_time:.3f}s, Success: {result.returncode == 0}")
                
            except subprocess.TimeoutExpired:
                loading_times[file_type] = {"error": "timeout"}
                print(f"    Timeout")
            except Exception as e:
                loading_times[file_type] = {"error": str(e)}
                print(f"    Error: {e}")
        
        # Cleanup test files
        self.cleanup_test_files(test_files)
        
        return loading_times
    
    def create_test_files(self) -> Dict[str, Path]:
        """Create test files for performance testing."""
        temp_dir = Path(tempfile.mkdtemp(prefix="vaitp_benchmark_"))
        test_files = {}
        
        # Create a small Excel file
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Data"
            
            # Add some test data
            for row in range(1, 101):  # 100 rows
                ws[f'A{row}'] = f"Item {row}"
                ws[f'B{row}'] = f"Description for item {row}"
                ws[f'C{row}'] = row * 10
            
            excel_file = temp_dir / "test_data.xlsx"
            wb.save(excel_file)
            test_files["excel"] = excel_file
            
        except ImportError:
            print("  Skipping Excel test (openpyxl not available)")
        
        # Create a SQLite database
        try:
            import sqlite3
            db_file = temp_dir / "test_data.db"
            conn = sqlite3.connect(db_file)
            cursor = conn.cursor()
            
            cursor.execute('''
                CREATE TABLE test_data (
                    id INTEGER PRIMARY KEY,
                    name TEXT,
                    description TEXT,
                    value INTEGER
                )
            ''')
            
            # Insert test data
            for i in range(1, 101):
                cursor.execute(
                    "INSERT INTO test_data (name, description, value) VALUES (?, ?, ?)",
                    (f"Item {i}", f"Description for item {i}", i * 10)
                )
            
            conn.commit()
            conn.close()
            test_files["sqlite"] = db_file
            
        except Exception as e:
            print(f"  Skipping SQLite test: {e}")
        
        return test_files
    
    def cleanup_test_files(self, test_files: Dict[str, Path]):
        """Clean up test files."""
        for file_path in test_files.values():
            try:
                if file_path.exists():
                    file_path.unlink()
                    
                # Remove temp directory if empty
                temp_dir = file_path.parent
                if temp_dir.exists() and not any(temp_dir.iterdir()):
                    temp_dir.rmdir()
                    
            except Exception as e:
                print(f"  Warning: Could not cleanup {file_path}: {e}")
    
    def benchmark_executable_size(self) -> Dict:
        """Analyze executable size and composition."""
        print("📏 Analyzing executable size...")
        
        if not self.executable_path.exists():
            return {"error": "Executable not found"}
        
        if self.executable_path.is_dir():
            # Calculate directory size (for .app bundles)
            total_size = 0
            file_count = 0
            
            for file_path in self.executable_path.rglob('*'):
                if file_path.is_file():
                    total_size += file_path.stat().st_size
                    file_count += 1
            
            results = {
                "total_size_mb": total_size / (1024 * 1024),
                "file_count": file_count,
                "type": "directory"
            }
        else:
            # Single file
            total_size = self.executable_path.stat().st_size
            results = {
                "total_size_mb": total_size / (1024 * 1024),
                "file_count": 1,
                "type": "file"
            }
        
        print(f"  Size: {results['total_size_mb']:.1f} MB")
        print(f"  Files: {results['file_count']}")
        
        return results
    
    def run_all_benchmarks(self) -> Dict:
        """Run all performance benchmarks."""
        print(f"🏁 Running performance benchmarks for: {self.executable_path}")
        print(f"Platform: {self.system} ({platform.machine()})")
        print("-" * 60)
        
        # Run benchmarks
        self.results["benchmarks"]["startup_time"] = self.benchmark_startup_time()
        self.results["benchmarks"]["memory_usage"] = self.benchmark_memory_usage()
        self.results["benchmarks"]["file_loading"] = self.benchmark_file_loading()
        self.results["benchmarks"]["executable_size"] = self.benchmark_executable_size()
        
        return self.results
    
    def print_summary(self):
        """Print benchmark summary."""
        print("\n" + "=" * 60)
        print("PERFORMANCE BENCHMARK SUMMARY")
        print("=" * 60)
        
        benchmarks = self.results["benchmarks"]
        
        # Startup time summary
        if "startup_time" in benchmarks and "mean_time" in benchmarks["startup_time"]:
            startup = benchmarks["startup_time"]
            print(f"🚀 Startup Time: {startup['mean_time']:.3f}s (±{startup['std_dev']:.3f}s)")
        
        # Memory usage summary
        if "memory_usage" in benchmarks and "memory" in benchmarks["memory_usage"]:
            memory = benchmarks["memory_usage"]["memory"]
            print(f"🧠 Memory Usage: {memory['mean_mb']:.1f} MB (peak: {memory['max_mb']:.1f} MB)")
        
        # Executable size summary
        if "executable_size" in benchmarks:
            size = benchmarks["executable_size"]
            print(f"📏 Executable Size: {size['total_size_mb']:.1f} MB")
        
        # Performance rating
        self.calculate_performance_rating()
    
    def calculate_performance_rating(self):
        """Calculate overall performance rating."""
        benchmarks = self.results["benchmarks"]
        score = 100  # Start with perfect score
        
        # Startup time penalty
        if "startup_time" in benchmarks and "mean_time" in benchmarks["startup_time"]:
            startup_time = benchmarks["startup_time"]["mean_time"]
            if startup_time > 5.0:
                score -= 30
            elif startup_time > 3.0:
                score -= 20
            elif startup_time > 2.0:
                score -= 10
        
        # Memory usage penalty
        if "memory_usage" in benchmarks and "memory" in benchmarks["memory_usage"]:
            memory_mb = benchmarks["memory_usage"]["memory"]["mean_mb"]
            if memory_mb > 200:
                score -= 25
            elif memory_mb > 100:
                score -= 15
            elif memory_mb > 50:
                score -= 5
        
        # Size penalty
        if "executable_size" in benchmarks:
            size_mb = benchmarks["executable_size"]["total_size_mb"]
            if size_mb > 200:
                score -= 20
            elif size_mb > 100:
                score -= 10
            elif size_mb > 50:
                score -= 5
        
        score = max(0, score)  # Don't go below 0
        
        if score >= 90:
            rating = "Excellent"
            emoji = "🏆"
        elif score >= 80:
            rating = "Good"
            emoji = "👍"
        elif score >= 70:
            rating = "Fair"
            emoji = "👌"
        elif score >= 60:
            rating = "Poor"
            emoji = "👎"
        else:
            rating = "Very Poor"
            emoji = "💥"
        
        print(f"\n{emoji} Overall Performance Rating: {rating} ({score}/100)")
        
        self.results["performance_rating"] = {
            "score": score,
            "rating": rating
        }
    
    def save_results(self, output_file: Path):
        """Save benchmark results to JSON file."""
        with open(output_file, 'w') as f:
            json.dump(self.results, f, indent=2)
        
        print(f"📄 Benchmark results saved to: {output_file}")


def find_executable(dist_dir: Path) -> Optional[Path]:
    """Find the executable in the dist directory."""
    if not dist_dir.exists():
        return None
    
    system = platform.system().lower()
    
    # Look for platform-specific executables
    for item in dist_dir.iterdir():
        if system == 'windows' and item.suffix == '.exe':
            return item
        elif system == 'darwin' and item.suffix == '.app':
            return item
        elif system == 'linux' and item.is_file() and not item.suffix and os.access(item, os.X_OK):
            return item
    
    return None


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Benchmark VAITP-Auditor GUI executable performance",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python benchmark_executable.py                           # Auto-find executable in dist/
  python benchmark_executable.py --executable dist/app.exe # Benchmark specific executable
  python benchmark_executable.py --output benchmark.json  # Save results to file
        """
    )
    
    parser.add_argument("--executable", type=Path, help="Path to executable to benchmark")
    parser.add_argument("--dist-dir", type=Path, default="dist", help="Distribution directory to search")
    parser.add_argument("--output", type=Path, help="Output file for benchmark results (JSON)")
    parser.add_argument("--iterations", type=int, default=10, help="Number of startup time iterations")
    parser.add_argument("--duration", type=int, default=10, help="Memory monitoring duration (seconds)")
    
    args = parser.parse_args()
    
    # Find executable
    if args.executable:
        executable_path = args.executable
    else:
        project_root = Path(__file__).parent.parent
        dist_dir = project_root / args.dist_dir
        executable_path = find_executable(dist_dir)
        
        if not executable_path:
            print(f"❌ No executable found in {dist_dir}")
            sys.exit(1)
    
    if not executable_path.exists():
        print(f"❌ Executable not found: {executable_path}")
        sys.exit(1)
    
    # Run benchmarks
    benchmark = PerformanceBenchmark(executable_path)
    results = benchmark.run_all_benchmarks()
    benchmark.print_summary()
    
    # Save results if requested
    if args.output:
        benchmark.save_results(args.output)
    
    print("\n🎯 Benchmark completed!")


if __name__ == "__main__":
    main()