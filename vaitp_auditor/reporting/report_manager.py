"""
Report manager for output file generation and management.
"""

import csv
import os
import tempfile
import threading
import time
import errno
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any
from ..core.models import ReviewResult

# Handle platform-specific file locking
try:
    import fcntl
    FCNTL_AVAILABLE = True
except ImportError:
    FCNTL_AVAILABLE = False
    # On Windows, we'll use a different approach

try:
    import msvcrt
    MSVCRT_AVAILABLE = True
except ImportError:
    MSVCRT_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


class ReportManager:
    """
    Manages output file generation with atomic writes.
    
    Handles both Excel and CSV formats with proper data integrity.
    Supports atomic result appending and undo functionality.
    """

    def __init__(self):
        """Initialize the report manager."""
        self._current_session_id: Optional[str] = None
        self._output_file_path: Optional[str] = None
        self._temp_file_path: Optional[str] = None
        self._output_format: str = 'excel'
        self._lock = threading.Lock()
        self._review_data: List[Dict[str, Any]] = []
        self._last_review_id: Optional[int] = None
        self._manual_verification_stats: Dict[str, int] = {
            'successful_injections': 0,
            'unsuccessful_injections': 0,
            'total_manual_verifications': 0
        }


    def initialize_report(self, session_id: str, output_format: str = 'excel') -> None:
        """
        Initialize a new report file for the session.
        
        Args:
            session_id: Unique session identifier.
            output_format: Output format ('excel' or 'csv').
            
        Raises:
            ValueError: If output_format is invalid or pandas not available for Excel.
            OSError: If unable to create output directory or temp file.
        """
        if output_format not in ['excel', 'csv']:
            raise ValueError(f"Invalid output format: {output_format}. Must be 'excel' or 'csv'")
        
        if output_format == 'excel' and not PANDAS_AVAILABLE:
            raise ValueError("Excel output requires pandas library. Please install pandas or use CSV format.")
        
        with self._lock:
            self._current_session_id = session_id
            self._output_format = output_format
            self._review_data = []
            self._last_review_id = None
            self._manual_verification_stats = {
                'successful_injections': 0,
                'unsuccessful_injections': 0,
                'total_manual_verifications': 0
            }
            
            # Create output directory if it doesn't exist
            output_dir = Path("reports")
            try:
                output_dir.mkdir(exist_ok=True)
            except PermissionError:
                # Try alternative locations if reports directory is not writable
                alternative_dirs = [
                    Path.home() / "vaitp_reports",
                    Path.cwd() / "temp_reports",
                    Path("/tmp") / "vaitp_reports" if os.name != 'nt' else Path.cwd() / "temp_reports"
                ]
                
                output_dir = None
                for alt_dir in alternative_dirs:
                    try:
                        alt_dir.mkdir(exist_ok=True)
                        output_dir = alt_dir
                        print(f"Warning: Using alternative output directory: {output_dir}")
                        break
                    except (PermissionError, OSError):
                        continue
                
                if output_dir is None:
                    raise OSError("Unable to create output directory. Please check permissions or specify a writable directory.")
            
            # Generate output file path with timestamp
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            file_extension = 'xlsx' if output_format == 'excel' else 'csv'
            filename = f"{session_id}_{timestamp}.{file_extension}"
            self._output_file_path = output_dir / filename
            
            # Create temporary file for atomic operations
            temp_dir = output_dir / "temp"
            try:
                temp_dir.mkdir(exist_ok=True)
                temp_fd, self._temp_file_path = tempfile.mkstemp(
                    suffix=f".{file_extension}",
                    prefix=f"{session_id}_temp_",
                    dir=temp_dir
                )
                os.close(temp_fd)  # Close the file descriptor, we'll manage the file ourselves
            except (PermissionError, OSError) as e:
                raise OSError(f"Unable to create temporary file in {temp_dir}. Error: {e}. "
                            f"Please ensure the directory is writable or try a different location.")
            
            # Initialize the file with headers
            self._write_headers()

    def resume_report(self, session_id: str, existing_file_path: str, output_format: str = 'excel') -> None:
        """
        Resume an existing report file for the session.
        
        Args:
            session_id: Unique session identifier.
            existing_file_path: Path to the existing report file to resume.
            output_format: Output format ('excel' or 'csv').
            
        Raises:
            ValueError: If output_format is invalid or pandas not available for Excel.
            OSError: If unable to access existing file or create temp file.
        """
        if output_format not in ['excel', 'csv']:
            raise ValueError(f"Invalid output format: {output_format}. Must be 'excel' or 'csv'")
        
        if output_format == 'excel' and not PANDAS_AVAILABLE:
            raise ValueError("Excel output requires pandas library. Please install pandas or use CSV format.")
        
        existing_path = Path(existing_file_path)
        if not existing_path.exists():
            raise ValueError(f"Existing report file not found: {existing_file_path}")
        
        with self._lock:
            self._current_session_id = session_id
            self._output_format = output_format
            self._output_file_path = existing_path
            
            # Load existing data from the file
            self._load_existing_data()
            
            # Create temporary file for atomic operations
            temp_dir = existing_path.parent / "temp"
            try:
                temp_dir.mkdir(exist_ok=True)
                temp_fd, self._temp_file_path = tempfile.mkstemp(
                    suffix=f".{existing_path.suffix.lstrip('.')}",
                    prefix=f"{session_id}_temp_",
                    dir=temp_dir
                )
                os.close(temp_fd)  # Close the file descriptor, we'll manage the file ourselves
            except (PermissionError, OSError) as e:
                raise OSError(f"Unable to create temporary file in {temp_dir}. Error: {e}. "
                            f"Please ensure the directory is writable or try a different location.")
            
            # Write current data to temp file to ensure consistency
            self._write_data_to_temp_file()

    def _load_existing_data(self) -> None:
        """Load existing data from the report file."""
        try:
            if self._output_format == 'excel' and PANDAS_AVAILABLE:
                # Load from Excel file
                df = pd.read_excel(self._output_file_path, engine='openpyxl')
                self._review_data = df.to_dict('records')
            elif self._output_format == 'csv':
                # Load from CSV file
                import csv
                self._review_data = []
                with open(self._output_file_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    self._review_data = list(reader)
            
            # Update last review ID and stats
            if self._review_data:
                # Find the highest review_id
                review_ids = [int(row.get('review_id', 0)) for row in self._review_data if row.get('review_id')]
                self._last_review_id = max(review_ids) if review_ids else None
                
                # Recalculate manual verification stats
                self._recalculate_stats()
            else:
                self._last_review_id = None
                self._manual_verification_stats = {
                    'successful_injections': 0,
                    'unsuccessful_injections': 0,
                    'total_manual_verifications': 0
                }
            
            print(f"Resumed report with {len(self._review_data)} existing reviews")
            
        except Exception as e:
            # If loading fails, start with empty data but log the error
            print(f"Warning: Failed to load existing report data: {e}")
            self._review_data = []
            self._last_review_id = None
            self._manual_verification_stats = {
                'successful_injections': 0,
                'unsuccessful_injections': 0,
                'total_manual_verifications': 0
            }

    def _recalculate_stats(self) -> None:
        """Recalculate manual verification statistics from existing data."""
        self._manual_verification_stats = {
            'successful_injections': 0,
            'unsuccessful_injections': 0,
            'total_manual_verifications': len(self._review_data)
        }
        
        for row in self._review_data:
            verdict = row.get('reviewer_verdict', '')
            if verdict in ['Success', 'Partial Success']:
                self._manual_verification_stats['successful_injections'] += 1
            elif verdict in ['Failure - No Change', 'Invalid Code', 'Wrong Vulnerability']:
                self._manual_verification_stats['unsuccessful_injections'] += 1

    def append_review_result(self, result: ReviewResult) -> None:
        """
        Append a review result to the report file atomically.
        
        Args:
            result: The review result to append.
            
        Raises:
            ValueError: If report not initialized or result validation fails.
            OSError: If unable to write to file.
        """
        if not self._current_session_id or not self._temp_file_path:
            raise ValueError("Report not initialized. Call initialize_report() first.")
        
        # Validate the review result
        if not result.validate_integrity():
            raise ValueError("Invalid review result data")
        
        with self._lock:
            # Convert ReviewResult to dictionary for storage
            result_dict = {
                'review_id': result.review_id,
                'source_identifier': result.source_identifier,
                'experiment_name': result.experiment_name,
                'review_timestamp_utc': result.review_timestamp_utc.isoformat(),
                'reviewer_verdict': result.reviewer_verdict,
                'reviewer_comment': result.reviewer_comment,
                'time_to_review_seconds': result.time_to_review_seconds,
                'model_name': result.model_name or '',
                'prompting_strategy': result.prompting_strategy or '',
                'expected_code': result.expected_code or '',
                'generated_code': result.generated_code,
                'code_diff': result.code_diff
            }
            
            # Add to in-memory data
            self._review_data.append(result_dict)
            self._last_review_id = result.review_id
            
            # Update manual verification statistics
            self._update_manual_verification_stats(result.reviewer_verdict)
            
            # For Excel format, check compatibility and sanitize data if needed
            # Do not automatically switch formats - respect user's choice
            if self._output_format == 'excel':
                is_compatible, reason = self._check_excel_compatibility([result_dict])
                if not is_compatible:
                    print(f"Warning: Data not fully compatible with Excel format ({reason}).")
                    print("Applying data sanitization to make it Excel-compatible...")
                    # Continue with Excel format but with sanitized data
            
            # Write to temporary file atomically with locking
            success = self._write_data_to_temp_file_with_locking()
            if not success:
                # Fallback to basic write method
                try:
                    self._write_data_to_temp_file()
                except Exception as e:
                    # If both methods fail, remove the added data and raise
                    self._review_data.pop()
                    if self._review_data:
                        self._last_review_id = self._review_data[-1]['review_id']
                    else:
                        self._last_review_id = None
                    
                    # Provide more specific error message for Excel compatibility issues
                    error_msg = str(e)
                    if "cannot be used in worksheets" in error_msg or "openpyxl" in error_msg:
                        raise OSError(f"Failed to write review result to Excel file due to incompatible content. "
                                    f"The generated code contains characters or data that Excel cannot handle even after sanitization. "
                                    f"To continue this session, please restart and select CSV format in the wizard. "
                                    f"Original error: {error_msg}")
                    else:
                        raise OSError(f"Failed to write review result to file: {error_msg}")

    def get_last_review_id(self) -> Optional[int]:
        """
        Get the ID of the last review for undo functionality.
        
        Returns:
            Optional[int]: Last review ID or None if no reviews exist.
        """
        with self._lock:
            return self._last_review_id

    def remove_last_review(self) -> bool:
        """
        Remove the last review from the report (for undo functionality).
        
        Implements atomic undo operations with proper error recovery and
        file locking to handle concurrent access gracefully.
        
        Returns:
            bool: True if removal was successful, False otherwise.
        """
        with self._lock:
            if not self._review_data:
                return False
            
            # Create backup of current state for rollback
            backup_data = self._review_data.copy()
            backup_last_id = self._last_review_id
            
            try:
                # Remove the last review
                removed_review = self._review_data.pop()
                
                # Update manual verification statistics (reverse the count)
                self._reverse_manual_verification_stats(removed_review.get('reviewer_verdict', ''))
                
                # Update last review ID
                if self._review_data:
                    self._last_review_id = self._review_data[-1]['review_id']
                else:
                    self._last_review_id = None
                
                # Attempt to rewrite the temporary file with file locking
                success = self._write_data_to_temp_file_with_locking()
                
                if not success:
                    # Rollback on failure
                    self._review_data = backup_data
                    self._last_review_id = backup_last_id
                    return False
                
                return True
                
            except Exception as e:
                # Rollback on any exception
                self._review_data = backup_data
                self._last_review_id = backup_last_id
                # Log the error but don't raise to maintain graceful degradation
                print(f"Warning: Failed to remove last review due to error: {e}")
                return False

    def finalize_report(self, output_format: Optional[str] = None) -> str:
        """
        Finalize the report and return the output file path.
        
        Args:
            output_format: Final output format ('excel' or 'csv'). 
                          If None, uses the format from initialization.
            
        Returns:
            str: Path to the finalized report file.
            
        Raises:
            ValueError: If report not initialized or invalid format.
            OSError: If unable to finalize the file.
        """
        if not self._current_session_id or not self._temp_file_path:
            raise ValueError("Report not initialized. Call initialize_report() first.")
        
        # Always use the requested format or the initialized format
        final_format = output_format or self._output_format
        
        if final_format not in ['excel', 'csv']:
            raise ValueError(f"Invalid output format: {final_format}")
        
        if final_format == 'excel' and not PANDAS_AVAILABLE:
            raise ValueError("Excel output requires pandas library")
        
        with self._lock:
            try:
                # Update output file path if format changed
                if final_format != self._output_format:
                    file_extension = 'xlsx' if final_format == 'excel' else 'csv'
                    self._output_file_path = self._output_file_path.with_suffix(f'.{file_extension}')
                
                # Move temp file to final location
                if Path(self._temp_file_path).exists():
                    # If format conversion is needed, do it now
                    if final_format != self._output_format:
                        self._convert_format(self._temp_file_path, self._output_file_path, final_format)
                    else:
                        # Simple move
                        Path(self._temp_file_path).rename(self._output_file_path)
                else:
                    # Create final file directly if temp doesn't exist
                    self._write_final_file(final_format)
                
                # Add statistics in the requested format only
                if final_format == 'excel':
                    # Excel file with integrated statistics sheet
                    print("Creating Excel file with integrated statistics sheet...")
                    self._add_statistics_to_excel_report()
                elif final_format == 'csv':
                    # CSV file with separate CSV statistics file
                    print("Creating CSV file with separate statistics file...")
                    self._create_statistics_csv_file()
                
                # Clean up temp file if it still exists
                if Path(self._temp_file_path).exists():
                    Path(self._temp_file_path).unlink()
                
                return str(self._output_file_path)
                
            except Exception as e:
                raise OSError(f"Failed to finalize report: {e}")

    def _write_headers(self) -> None:
        """Write column headers to the temporary file."""
        headers = [
            'review_id',
            'source_identifier', 
            'experiment_name',
            'review_timestamp_utc',
            'reviewer_verdict',
            'reviewer_comment',
            'time_to_review_seconds',
            'model_name',
            'prompting_strategy',
            'expected_code',
            'generated_code',
            'code_diff'
        ]
        
        if self._output_format == 'csv':
            with open(self._temp_file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
        elif self._output_format == 'excel' and PANDAS_AVAILABLE:
            # Create empty DataFrame with headers for Excel
            df = pd.DataFrame(columns=headers)
            df.to_excel(self._temp_file_path, index=False, engine='openpyxl')

    def _write_data_to_temp_file(self) -> None:
        """Write all current data to the temporary file."""
        if self._output_format == 'csv':
            self._write_csv_data()
        elif self._output_format == 'excel' and PANDAS_AVAILABLE:
            self._write_excel_data()

    def _write_data_to_temp_file_with_locking(self) -> bool:
        """
        Write all current data to the temporary file with file locking.
        
        Implements proper file locking and permission error handling
        to support concurrent access and graceful degradation.
        
        Returns:
            bool: True if write was successful, False otherwise.
        """
        try:
            if self._output_format == 'csv':
                return self._write_csv_data_with_locking()
            elif self._output_format == 'excel' and PANDAS_AVAILABLE:
                return self._write_excel_data_with_locking()
            else:
                return False
        except (OSError, IOError, PermissionError) as e:
            print(f"Warning: Failed to write to temporary file: {e}")
            if e.errno == errno.EACCES:
                print("Suggestion: Check file permissions or try running with appropriate privileges")
            elif e.errno == errno.ENOSPC:
                print("Suggestion: Free up disk space and try again")
            elif e.errno == errno.EROFS:
                print("Suggestion: The file system is read-only, try a different output directory")
            return False
        except Exception as e:
            print(f"Warning: Unexpected error writing to temporary file: {e}")
            return False

    def _write_csv_data(self) -> None:
        """Write data to CSV format."""
        headers = [
            'review_id',
            'source_identifier', 
            'experiment_name',
            'review_timestamp_utc',
            'reviewer_verdict',
            'reviewer_comment',
            'time_to_review_seconds',
            'model_name',
            'prompting_strategy',
            'expected_code',
            'generated_code',
            'code_diff'
        ]
        
        with open(self._temp_file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            for row_data in self._review_data:
                row = [row_data.get(header, '') for header in headers]
                writer.writerow(row)

    def _write_excel_data(self) -> None:
        """Write data to Excel format."""
        if not PANDAS_AVAILABLE:
            raise ValueError("Pandas required for Excel output")
        
        # Sanitize data for Excel compatibility
        sanitized_data = self._sanitize_data_for_excel(self._review_data)
        
        df = pd.DataFrame(sanitized_data)
        
        # Ensure proper column order
        column_order = [
            'review_id',
            'source_identifier', 
            'experiment_name',
            'review_timestamp_utc',
            'reviewer_verdict',
            'reviewer_comment',
            'time_to_review_seconds',
            'model_name',
            'prompting_strategy',
            'expected_code',
            'generated_code',
            'code_diff'
        ]
        
        # Reorder columns and fill missing ones
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
        
        df = df[column_order]
        df.to_excel(self._temp_file_path, index=False, engine='openpyxl')

    def _convert_format(self, source_path: str, target_path: Path, target_format: str) -> None:
        """Convert between CSV and Excel formats."""
        if target_format == 'csv':
            # Convert Excel to CSV
            if PANDAS_AVAILABLE:
                df = pd.read_excel(source_path, engine='openpyxl')
                df.to_csv(target_path, index=False)
            else:
                raise ValueError("Pandas required for format conversion")
        elif target_format == 'excel':
            # Convert CSV to Excel
            if PANDAS_AVAILABLE:
                df = pd.read_csv(source_path)
                df.to_excel(target_path, index=False, engine='openpyxl')
            else:
                raise ValueError("Pandas required for Excel output")

    def _write_final_file(self, output_format: str) -> None:
        """Write final file directly from in-memory data."""
        if output_format == 'csv':
            self._write_csv_data_to_path(self._output_file_path)
        elif output_format == 'excel':
            self._write_excel_data_to_path(self._output_file_path)

    def _write_csv_data_to_path(self, file_path: Path) -> None:
        """Write CSV data to specified path."""
        headers = [
            'review_id',
            'source_identifier', 
            'experiment_name',
            'review_timestamp_utc',
            'reviewer_verdict',
            'reviewer_comment',
            'time_to_review_seconds',
            'model_name',
            'prompting_strategy',
            'expected_code',
            'generated_code',
            'code_diff'
        ]
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            
            for row_data in self._review_data:
                row = [row_data.get(header, '') for header in headers]
                writer.writerow(row)

    def _write_excel_data_to_path(self, file_path: Path) -> None:
        """Write Excel data to specified path."""
        if not PANDAS_AVAILABLE:
            raise ValueError("Pandas required for Excel output")
        
        # Sanitize data for Excel compatibility
        sanitized_data = self._sanitize_data_for_excel(self._review_data)
        
        df = pd.DataFrame(sanitized_data)
        
        # Ensure proper column order
        column_order = [
            'review_id',
            'source_identifier', 
            'experiment_name',
            'review_timestamp_utc',
            'reviewer_verdict',
            'reviewer_comment',
            'time_to_review_seconds',
            'model_name',
            'prompting_strategy',
            'expected_code',
            'generated_code',
            'code_diff'
        ]
        
        # Reorder columns and fill missing ones
        for col in column_order:
            if col not in df.columns:
                df[col] = ''
        
        df = df[column_order]
        df.to_excel(file_path, index=False, engine='openpyxl')

    def _write_csv_data_with_locking(self) -> bool:
        """
        Write CSV data to temporary file with file locking.
        
        Returns:
            bool: True if write was successful, False otherwise.
        """
        headers = [
            'review_id',
            'source_identifier', 
            'experiment_name',
            'review_timestamp_utc',
            'reviewer_verdict',
            'reviewer_comment',
            'time_to_review_seconds',
            'model_name',
            'prompting_strategy',
            'expected_code',
            'generated_code',
            'code_diff'
        ]
        
        max_retries = 3
        retry_delay = 0.1
        
        for attempt in range(max_retries):
            try:
                with open(self._temp_file_path, 'w', newline='', encoding='utf-8') as f:
                    # Attempt to acquire exclusive lock (platform-specific)
                    lock_acquired = self._acquire_file_lock(f)
                    if not lock_acquired:
                        if attempt < max_retries - 1:
                            time.sleep(retry_delay * (2 ** attempt))  # Exponential backoff
                            continue
                        else:
                            return False
                    
                    writer = csv.writer(f)
                    writer.writerow(headers)
                    
                    for row_data in self._review_data:
                        row = [row_data.get(header, '') for header in headers]
                        writer.writerow(row)
                    
                    # Lock is automatically released when file is closed
                    return True
                    
            except (OSError, IOError, PermissionError):
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (2 ** attempt))
                    continue
                else:
                    return False
        
        return False

    def _sanitize_data_for_excel(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Sanitize data for Excel compatibility by handling problematic characters and content.
        
        Args:
            data: List of dictionaries containing review data
            
        Returns:
            List[Dict[str, Any]]: Sanitized data safe for Excel
        """
        sanitized_data = []
        
        for row in data:
            sanitized_row = {}
            for key, value in row.items():
                if isinstance(value, str):
                    # Handle problematic characters and content for Excel
                    sanitized_value = value
                    
                    # Replace null bytes which Excel cannot handle
                    sanitized_value = sanitized_value.replace('\x00', '[NULL]')
                    
                    # Replace other control characters that might cause issues
                    for i in range(1, 32):  # Control characters except tab, newline, carriage return
                        if i not in [9, 10, 13]:  # Keep tab, LF, CR
                            sanitized_value = sanitized_value.replace(chr(i), f'[CTRL-{i}]')
                    
                    # Handle very long strings that might cause Excel issues
                    # Excel has a cell limit of 32,767 characters
                    if len(sanitized_value) > 32000:
                        sanitized_value = sanitized_value[:32000] + "\n[TRUNCATED: Content too long for Excel]"
                    
                    # Handle binary-like content that appears as text
                    if "b'" in sanitized_value and ("adshibe_desires" in sanitized_value or len(sanitized_value) > 1000):
                        # This looks like problematic binary content or very long repetitive content
                        # that might cause Excel issues, make it Excel-safe
                        try:
                            # Find binary literals and replace them
                            import re
                            binary_pattern = r"b'[^']*'"
                            matches = re.findall(binary_pattern, sanitized_value)
                            for match in matches:
                                # Replace binary literal with safe representation
                                safe_repr = f"BINARY_DATA({len(match)} chars)"
                                sanitized_value = sanitized_value.replace(match, safe_repr)
                        except:
                            # If regex processing fails, just truncate and mark as binary
                            sanitized_value = f"BINARY_DATA: {sanitized_value[:1000]}{'...[TRUNCATED]' if len(sanitized_value) > 1000 else ''}"
                    
                    sanitized_row[key] = sanitized_value
                else:
                    # Non-string values should be fine
                    sanitized_row[key] = value
            
            sanitized_data.append(sanitized_row)
        
        return sanitized_data

    def _check_excel_compatibility(self, data: List[Dict[str, Any]]) -> tuple[bool, str]:
        """
        Check if data is compatible with Excel format.
        
        Args:
            data: List of dictionaries containing review data
            
        Returns:
            tuple[bool, str]: (is_compatible, reason_if_not)
        """
        for row in data:
            for key, value in row.items():
                if isinstance(value, str):
                    # Check for null bytes
                    if '\x00' in value:
                        return False, f"Contains null bytes in field '{key}'"
                    
                    # Check for very long content
                    if len(value) > 32767:
                        return False, f"Content too long for Excel in field '{key}' ({len(value)} chars)"
                    
                    # Check for problematic binary patterns
                    if "b'" in value and len(value) > 1000:
                        return False, f"Contains large binary data in field '{key}'"
                    
                    # Check for control characters
                    for i in range(1, 32):
                        if i not in [9, 10, 13] and chr(i) in value:
                            return False, f"Contains control character (ASCII {i}) in field '{key}'"
        
        return True, ""

    def _write_excel_data_with_locking(self) -> bool:
        """
        Write Excel data to temporary file with file locking.
        
        Returns:
            bool: True if write was successful, False otherwise.
        """
        if not PANDAS_AVAILABLE:
            return False
        
        max_retries = 3
        retry_delay = 0.1
        
        for attempt in range(max_retries):
            try:
                # Sanitize data for Excel compatibility
                sanitized_data = self._sanitize_data_for_excel(self._review_data)
                
                # Create DataFrame
                df = pd.DataFrame(sanitized_data)
                
                # Ensure proper column order
                column_order = [
                    'review_id',
                    'source_identifier', 
                    'experiment_name',
                    'review_timestamp_utc',
                    'reviewer_verdict',
                    'reviewer_comment',
                    'time_to_review_seconds',
                    'model_name',
                    'prompting_strategy',
                    'expected_code',
                    'generated_code',
                    'code_diff'
                ]
                
                # Reorder columns and fill missing ones
                for col in column_order:
                    if col not in df.columns:
                        df[col] = ''
                
                df = df[column_order]
                
                # Write to temporary file first, then move to avoid corruption
                # Use .xlsx extension for pandas compatibility
                temp_excel_path = f"{self._temp_file_path}.writing.xlsx"
                
                try:
                    df.to_excel(temp_excel_path, index=False, engine='openpyxl')
                    
                    # Remove existing temp file if it exists
                    if os.path.exists(self._temp_file_path):
                        os.unlink(self._temp_file_path)
                    
                    # Atomic move to final temp location
                    os.rename(temp_excel_path, self._temp_file_path)
                    return True
                    
                except (OSError, IOError, PermissionError):
                    # Clean up temporary file if it exists
                    if os.path.exists(temp_excel_path):
                        try:
                            os.unlink(temp_excel_path)
                        except:
                            pass
                    
                    if attempt < max_retries - 1:
                        time.sleep(retry_delay * (2 ** attempt))
                        continue
                    else:
                        return False
                        
            except Exception:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (2 ** attempt))
                    continue
                else:
                    return False
        
        return False

    def _acquire_file_lock(self, file_obj) -> bool:
        """
        Acquire an exclusive file lock in a cross-platform manner.
        
        Args:
            file_obj: Open file object to lock.
            
        Returns:
            bool: True if lock was acquired, False otherwise.
        """
        try:
            if FCNTL_AVAILABLE:
                # Unix/Linux/macOS
                fcntl.flock(file_obj.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                return True
            elif MSVCRT_AVAILABLE:
                # Windows
                msvcrt.locking(file_obj.fileno(), msvcrt.LK_NBLCK, 1)
                return True
            else:
                # Fallback: no locking available, but don't fail
                return True
        except (OSError, IOError):
            return False
    
    def _update_manual_verification_stats(self, verdict: str) -> None:
        """
        Update manual verification statistics based on verdict.
        
        Args:
            verdict: The reviewer verdict
        """
        # Increment total manual verifications
        self._manual_verification_stats['total_manual_verifications'] += 1
        
        # Categorize verdict as successful or unsuccessful injection
        if verdict in ['Success', 'Partial Success']:
            self._manual_verification_stats['successful_injections'] += 1
        elif verdict in ['Failure - No Change', 'Invalid Code', 'Wrong Vulnerability']:
            self._manual_verification_stats['unsuccessful_injections'] += 1
        # Note: 'Flag NOT Vulnerable Expected' and other verdicts don't fit either category
    
    def _reverse_manual_verification_stats(self, verdict: str) -> None:
        """
        Reverse manual verification statistics when undoing a review.
        
        Args:
            verdict: The reviewer verdict to reverse
        """
        # Decrement total manual verifications
        if self._manual_verification_stats['total_manual_verifications'] > 0:
            self._manual_verification_stats['total_manual_verifications'] -= 1
        
        # Reverse categorization
        if verdict in ['Success', 'Partial Success']:
            if self._manual_verification_stats['successful_injections'] > 0:
                self._manual_verification_stats['successful_injections'] -= 1
        elif verdict in ['Failure - No Change', 'Invalid Code', 'Wrong Vulnerability']:
            if self._manual_verification_stats['unsuccessful_injections'] > 0:
                self._manual_verification_stats['unsuccessful_injections'] -= 1
    
    def _add_statistics_to_excel_report(self) -> None:
        """
        Add a comprehensive statistics sheet to the Excel report.
        """
        if not PANDAS_AVAILABLE:
            return
        
        try:
            # Calculate comprehensive statistics
            stats = self._calculate_comprehensive_statistics()
            
            # Use openpyxl directly for better control
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Load the existing workbook
            workbook = load_workbook(self._output_file_path)
            
            # Create a new worksheet for statistics
            if 'Statistics' in workbook.sheetnames:
                del workbook['Statistics']
            stats_sheet = workbook.create_sheet('Statistics')
            
            # Create main statistics data
            stats_data = [
                ['Metric', 'Count', 'Percentage'],
                ['Total Reviews Completed', str(stats['total_reviews']), '100.0%'],
                ['', '', ''],
                ['VERDICT BREAKDOWN', '', ''],
                ['Success', str(stats['verdict_counts']['Success']), f"{stats['verdict_percentages']['Success']:.1f}%"],
                ['Partial Success', str(stats['verdict_counts']['Partial Success']), f"{stats['verdict_percentages']['Partial Success']:.1f}%"],
                ['Failure - No Change', str(stats['verdict_counts']['Failure - No Change']), f"{stats['verdict_percentages']['Failure - No Change']:.1f}%"],
                ['Invalid Code', str(stats['verdict_counts']['Invalid Code']), f"{stats['verdict_percentages']['Invalid Code']:.1f}%"],
                ['Wrong Vulnerability', str(stats['verdict_counts']['Wrong Vulnerability']), f"{stats['verdict_percentages']['Wrong Vulnerability']:.1f}%"],
                ['Flag NOT Vulnerable Expected', str(stats['verdict_counts']['Flag NOT Vulnerable Expected']), f"{stats['verdict_percentages']['Flag NOT Vulnerable Expected']:.1f}%"],
                ['Other/Custom', str(stats['verdict_counts']['Other']), f"{stats['verdict_percentages']['Other']:.1f}%"],
                ['', '', ''],
                ['SUMMARY CATEGORIES', '', ''],
                ['Successful Outcomes', str(stats['successful_outcomes']), f"{stats['successful_percentage']:.1f}%"],
                ['Failed Outcomes', str(stats['failed_outcomes']), f"{stats['failed_percentage']:.1f}%"],
                ['Generated Code Classified as Wrong Vulnerability', str(stats['verdict_counts']['Wrong Vulnerability']), f"{stats['verdict_percentages']['Wrong Vulnerability']:.1f}%"],
                ['', '', ''],
                ['PERFORMANCE METRICS', '', ''],
                ['Average Review Time (seconds)', f"{stats['avg_review_time']:.2f}", ''],
                ['Median Review Time (seconds)', f"{stats['median_review_time']:.2f}", ''],
                ['Total Review Time (minutes)', f"{stats['total_review_time_minutes']:.1f}", ''],
                ['', '', ''],
                ['MODEL BREAKDOWN', '', ''],
            ]
            
            # Add model statistics if available
            for model, count in stats['model_counts'].items():
                percentage = (count / max(1, stats['total_reviews'])) * 100
                stats_data.append([f"Model: {model}", str(count), f"{percentage:.1f}%"])
            
            if not stats['model_counts']:
                stats_data.append(['No model information available', '', ''])
            
            # Write data to worksheet row by row to avoid formula interpretation
            for row_idx, row_data in enumerate(stats_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    cell = stats_sheet.cell(row=row_idx, column=col_idx)
                    # Explicitly set as text to prevent formula interpretation
                    cell.value = str(cell_value) if cell_value else ''
            
            # Save the workbook
            workbook.save(self._output_file_path)
            workbook.close()
                
        except Exception as e:
            # Statistics are optional, don't fail the entire report
            print(f"Warning: Failed to add statistics sheet to Excel report: {e}")
            # Fallback to pandas method if openpyxl fails
            try:
                self._add_statistics_to_excel_report_pandas_fallback(stats)
            except Exception as fallback_error:
                print(f"Warning: Fallback method also failed: {fallback_error}")
    
    def _add_statistics_to_excel_report_pandas_fallback(self, stats: Dict[str, Any]) -> None:
        """
        Fallback method using pandas for Excel statistics.
        """
        with pd.ExcelWriter(self._output_file_path, engine='openpyxl', mode='a') as writer:
            # Create main statistics data
            stats_data = [
                ['Metric', 'Count', 'Percentage'],
                ['Total Reviews Completed', stats['total_reviews'], '100.0%'],
                ['', '', ''],
                ['VERDICT BREAKDOWN', '', ''],
                ['Success', stats['verdict_counts']['Success'], f"{stats['verdict_percentages']['Success']:.1f}%"],
                ['Partial Success', stats['verdict_counts']['Partial Success'], f"{stats['verdict_percentages']['Partial Success']:.1f}%"],
                ['Failure - No Change', stats['verdict_counts']['Failure - No Change'], f"{stats['verdict_percentages']['Failure - No Change']:.1f}%"],
                ['Invalid Code', stats['verdict_counts']['Invalid Code'], f"{stats['verdict_percentages']['Invalid Code']:.1f}%"],
                ['Wrong Vulnerability', stats['verdict_counts']['Wrong Vulnerability'], f"{stats['verdict_percentages']['Wrong Vulnerability']:.1f}%"],
                ['Flag NOT Vulnerable Expected', stats['verdict_counts']['Flag NOT Vulnerable Expected'], f"{stats['verdict_percentages']['Flag NOT Vulnerable Expected']:.1f}%"],
                ['Other/Custom', stats['verdict_counts']['Other'], f"{stats['verdict_percentages']['Other']:.1f}%"],
                ['', '', ''],
                ['SUMMARY CATEGORIES', '', ''],
                ['Successful Outcomes', stats['successful_outcomes'], f"{stats['successful_percentage']:.1f}%"],
                ['Failed Outcomes', stats['failed_outcomes'], f"{stats['failed_percentage']:.1f}%"],
                ['Generated Code Classified as Wrong Vulnerability', stats['verdict_counts']['Wrong Vulnerability'], f"{stats['verdict_percentages']['Wrong Vulnerability']:.1f}%"],
                ['', '', ''],
                ['PERFORMANCE METRICS', '', ''],
                ['Average Review Time (seconds)', f"{stats['avg_review_time']:.2f}", ''],
                ['Median Review Time (seconds)', f"{stats['median_review_time']:.2f}", ''],
                ['Total Review Time (minutes)', f"{stats['total_review_time_minutes']:.1f}", ''],
                ['', '', ''],
                ['MODEL BREAKDOWN', '', ''],
            ]
            
            # Add model statistics if available
            for model, count in stats['model_counts'].items():
                percentage = (count / max(1, stats['total_reviews'])) * 100
                stats_data.append([f"Model: {model}", count, f"{percentage:.1f}%"])
            
            if not stats['model_counts']:
                stats_data.append(['No model information available', '', ''])
            
            # Create DataFrame for statistics
            stats_df = pd.DataFrame(stats_data[1:], columns=stats_data[0])
            
            # Ensure all data is treated as text to avoid formula interpretation
            for col in stats_df.columns:
                stats_df[col] = stats_df[col].astype(str)
            
            # Write to new sheet
            stats_df.to_excel(writer, sheet_name='Statistics', index=False)
    
    def _create_statistics_csv_file(self) -> None:
        """
        Create a separate CSV file with comprehensive statistics.
        """
        try:
            # Create statistics file path
            stats_file_path = self._output_file_path.with_name(
                self._output_file_path.stem + '_statistics.csv'
            )
            
            # Calculate comprehensive statistics
            stats = self._calculate_comprehensive_statistics()
            
            # Create statistics data
            stats_data = [
                ['Metric', 'Count', 'Percentage'],
                ['Total Reviews Completed', stats['total_reviews'], '100.0%'],
                ['', '', ''],
                ['VERDICT BREAKDOWN', '', ''],
                ['Success', stats['verdict_counts']['Success'], f"{stats['verdict_percentages']['Success']:.1f}%"],
                ['Partial Success', stats['verdict_counts']['Partial Success'], f"{stats['verdict_percentages']['Partial Success']:.1f}%"],
                ['Failure - No Change', stats['verdict_counts']['Failure - No Change'], f"{stats['verdict_percentages']['Failure - No Change']:.1f}%"],
                ['Invalid Code', stats['verdict_counts']['Invalid Code'], f"{stats['verdict_percentages']['Invalid Code']:.1f}%"],
                ['Wrong Vulnerability', stats['verdict_counts']['Wrong Vulnerability'], f"{stats['verdict_percentages']['Wrong Vulnerability']:.1f}%"],
                ['Flag NOT Vulnerable Expected', stats['verdict_counts']['Flag NOT Vulnerable Expected'], f"{stats['verdict_percentages']['Flag NOT Vulnerable Expected']:.1f}%"],
                ['Other/Custom', stats['verdict_counts']['Other'], f"{stats['verdict_percentages']['Other']:.1f}%"],
                ['', '', ''],
                ['SUMMARY CATEGORIES', '', ''],
                ['Successful Outcomes', stats['successful_outcomes'], f"{stats['successful_percentage']:.1f}%"],
                ['Failed Outcomes', stats['failed_outcomes'], f"{stats['failed_percentage']:.1f}%"],
                ['Generated Code Classified as Wrong Vulnerability', stats['verdict_counts']['Wrong Vulnerability'], f"{stats['verdict_percentages']['Wrong Vulnerability']:.1f}%"],
                ['', '', ''],
                ['PERFORMANCE METRICS', '', ''],
                ['Average Review Time (seconds)', f"{stats['avg_review_time']:.2f}", ''],
                ['Median Review Time (seconds)', f"{stats['median_review_time']:.2f}", ''],
                ['Total Review Time (minutes)', f"{stats['total_review_time_minutes']:.1f}", ''],
                ['', '', ''],
                ['MODEL BREAKDOWN', '', ''],
            ]
            
            # Add model statistics if available
            for model, count in stats['model_counts'].items():
                percentage = (count / max(1, stats['total_reviews'])) * 100
                stats_data.append([f"Model: {model}", count, f"{percentage:.1f}%"])
            
            if not stats['model_counts']:
                stats_data.append(['No model information available', '', ''])
            
            # Write statistics to CSV
            with open(stats_file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(stats_data)
                
        except Exception as e:
            # Statistics are optional, don't fail the entire report
            print(f"Warning: Failed to create statistics CSV file: {e}")
    
    def _calculate_comprehensive_statistics(self) -> Dict[str, Any]:
        """
        Calculate comprehensive statistics from review data.
        
        Returns:
            Dictionary containing detailed statistics
        """
        if not self._review_data:
            return self._get_empty_statistics()
        
        total_reviews = len(self._review_data)
        
        # Count verdicts
        verdict_counts = {
            'Success': 0,
            'Partial Success': 0,
            'Failure - No Change': 0,
            'Invalid Code': 0,
            'Wrong Vulnerability': 0,
            'Flag NOT Vulnerable Expected': 0,
            'Other': 0
        }
        
        # Count models
        model_counts = {}
        
        # Calculate review times
        review_times = []
        
        for row in self._review_data:
            verdict = row.get('reviewer_verdict', '')
            
            # Count verdicts
            if verdict in verdict_counts:
                verdict_counts[verdict] += 1
            else:
                verdict_counts['Other'] += 1
            
            # Count models
            model = row.get('model_name', 'Unknown')
            if model and model.strip():
                model_counts[model] = model_counts.get(model, 0) + 1
            else:
                model_counts['Unknown'] = model_counts.get('Unknown', 0) + 1
            
            # Collect review times
            try:
                time_val = float(row.get('time_to_review_seconds', 0))
                if time_val > 0:
                    review_times.append(time_val)
            except (ValueError, TypeError):
                pass
        
        # Calculate percentages
        verdict_percentages = {}
        for verdict, count in verdict_counts.items():
            verdict_percentages[verdict] = (count / total_reviews) * 100
        
        # Calculate success/failure categories
        successful_outcomes = verdict_counts['Success'] + verdict_counts['Partial Success']
        failed_outcomes = (verdict_counts['Failure - No Change'] + 
                          verdict_counts['Invalid Code'] + 
                          verdict_counts['Wrong Vulnerability'])
        
        successful_percentage = (successful_outcomes / total_reviews) * 100
        failed_percentage = (failed_outcomes / total_reviews) * 100
        
        # Calculate time statistics
        avg_review_time = sum(review_times) / len(review_times) if review_times else 0.0
        median_review_time = 0.0
        if review_times:
            sorted_times = sorted(review_times)
            n = len(sorted_times)
            if n % 2 == 0:
                median_review_time = (sorted_times[n//2 - 1] + sorted_times[n//2]) / 2
            else:
                median_review_time = sorted_times[n//2]
        
        total_review_time_minutes = sum(review_times) / 60 if review_times else 0.0
        
        return {
            'total_reviews': total_reviews,
            'verdict_counts': verdict_counts,
            'verdict_percentages': verdict_percentages,
            'successful_outcomes': successful_outcomes,
            'failed_outcomes': failed_outcomes,
            'successful_percentage': successful_percentage,
            'failed_percentage': failed_percentage,
            'model_counts': model_counts,
            'avg_review_time': avg_review_time,
            'median_review_time': median_review_time,
            'total_review_time_minutes': total_review_time_minutes
        }
    
    def _get_empty_statistics(self) -> Dict[str, Any]:
        """
        Return empty statistics structure when no data is available.
        
        Returns:
            Dictionary with zero values for all statistics
        """
        return {
            'total_reviews': 0,
            'verdict_counts': {
                'Success': 0,
                'Partial Success': 0,
                'Failure - No Change': 0,
                'Invalid Code': 0,
                'Wrong Vulnerability': 0,
                'Flag NOT Vulnerable Expected': 0,
                'Other': 0
            },
            'verdict_percentages': {
                'Success': 0.0,
                'Partial Success': 0.0,
                'Failure - No Change': 0.0,
                'Invalid Code': 0.0,
                'Wrong Vulnerability': 0.0,
                'Flag NOT Vulnerable Expected': 0.0,
                'Other': 0.0
            },
            'successful_outcomes': 0,
            'failed_outcomes': 0,
            'successful_percentage': 0.0,
            'failed_percentage': 0.0,
            'model_counts': {},
            'avg_review_time': 0.0,
            'median_review_time': 0.0,
            'total_review_time_minutes': 0.0
        }

    def get_manual_verification_statistics(self) -> Dict[str, Any]:
        """
        Get current manual verification statistics.
        
        Returns:
            Dictionary containing manual verification statistics
        """
        total = self._manual_verification_stats['total_manual_verifications']
        successful = self._manual_verification_stats['successful_injections']
        unsuccessful = self._manual_verification_stats['unsuccessful_injections']
        
        return {
            'total_manual_verifications': total,
            'successful_injections': successful,
            'unsuccessful_injections': unsuccessful,
            'successful_percentage': (successful / max(1, total)) * 100,
            'unsuccessful_percentage': (unsuccessful / max(1, total)) * 100,
            'average_review_time_seconds': sum(float(row.get('time_to_review_seconds', 0)) for row in self._review_data) / max(1, len(self._review_data)) if self._review_data else 0.0
        }
    
    def get_comprehensive_statistics(self) -> Dict[str, Any]:
        """
        Get comprehensive statistics for external use.
        
        Returns:
            Dictionary containing detailed statistics
        """
        return self._calculate_comprehensive_statistics()